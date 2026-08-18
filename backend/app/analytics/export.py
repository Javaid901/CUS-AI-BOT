"""
backend/app/analytics/export.py

Export analytics data in multiple formats.

Supported formats: CSV, Excel (XLSX), PDF, JSON

Uses in-memory generation with appropriate MIME types.
"""

from __future__ import annotations

import csv
import io
import json
from typing import Any


def export_csv(data: list[dict[str, Any]]) -> str:
    """Export analytics data as CSV string."""
    if not data:
        return ""
    output = io.StringIO()
    writer = csv.DictWriter(output, fieldnames=data[0].keys())
    writer.writeheader()
    writer.writerows(data)
    return output.getvalue()


def export_json(data: Any, pretty: bool = True) -> str:
    """Export analytics data as JSON string."""
    indent = 2 if pretty else None
    return json.dumps(data, indent=indent, default=str)


def export_excel(data: list[dict[str, Any]], sheet_name: str = "Analytics") -> bytes:
    """Export analytics data as Excel XLSX bytes.

    Uses the openpyxl library if available, falls back to CSV wrapped in
    a simple XML spreadsheet.
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
    except ImportError:
        return _export_xlsx_fallback(data, sheet_name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    if not data:
        ws["A1"] = "No data"
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    # Header row
    headers = list(data[0].keys())
    header_fill = PatternFill(start_color="0F5132", end_color="0F5132", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font

    # Data rows
    for row_idx, row in enumerate(data, 2):
        for col_idx, h in enumerate(headers, 1):
            val = row.get(h)
            if isinstance(val, (dict, list)):
                val = json.dumps(val)
            ws.cell(row=row_idx, column=col_idx, value=val)

    # Auto-width
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 2, 60)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _export_xlsx_fallback(data: list[dict[str, Any]], sheet_name: str) -> bytes:
    """Fallback XLSX-like export using basic XML (for minimal dependency mode)."""
    rows = []
    if not data:
        rows.append("<row><cell>No data</cell></row>")
    else:
        headers = list(data[0].keys())
        rows.append("<row>" + "".join(f"<cell><b>{h}</b></cell>" for h in headers) + "</row>")
        for row in data:
            cells = ""
            for h in headers:
                val = row.get(h, "")
                if isinstance(val, (dict, list)):
                    val = json.dumps(val)
                cells += f"<cell>{val}</cell>"
            rows.append(f"<row>{cells}</row>")

    xml = f"""<?xml version="1.0"?>
<?mso-application progid="Excel.Sheet"?>
<Workbook xmlns="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:o="urn:schemas-microsoft-com:office:office"
 xmlns:x="urn:schemas-microsoft-com:office:excel"
 xmlns:ss="urn:schemas-microsoft-com:office:spreadsheet"
 xmlns:html="http://www.w3.org/TR/REC-html40">
<Worksheet ss:Name="{sheet_name}">
<Table>
{chr(10).join(rows)}
</Table>
</Worksheet>
</Workbook>"""
    return xml.encode("utf-8")


def export_pdf(title: str, sections: list[dict[str, Any]]) -> bytes:
    """Export analytics as a PDF report.

    Uses reportlab if available, falls back to simple HTML-to-PDF.
    Returns PDF bytes.
    """
    try:
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import A4
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        from reportlab.platypus import (
            Paragraph,
            SimpleDocTemplate,
            Spacer,
            Table,
            TableStyle,
        )
    except ImportError:
        return _export_pdf_fallback(title, sections)

    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, title=title)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph(title, styles["Title"]))
    story.append(Spacer(1, 0.25 * inch))

    for section in sections:
        story.append(Paragraph(section.get("title", ""), styles["Heading2"]))
        story.append(Spacer(1, 0.1 * inch))

        table_data = section.get("data", [])
        if table_data and isinstance(table_data, list):
            if table_data and isinstance(table_data[0], dict):
                headers = list(table_data[0].keys())
                rows = [[h.capitalize() for h in headers]]
                for row in table_data:
                    rows.append([str(row.get(h, "")) for h in headers])
                t = Table(rows, hAlign="LEFT")
                t.setStyle(TableStyle([
                    ("BACKGROUND", (0, 0), (-1, 0), colors.Color(15 / 255, 81 / 255, 50 / 255)),
                    ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                    ("FONTSIZE", (0, 0), (-1, -1), 9),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
                    ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                ]))
                story.append(t)
            else:
                story.append(Paragraph(str(table_data), styles["Normal"]))
        story.append(Spacer(1, 0.15 * inch))

    doc.build(story)
    return buf.getvalue()


def _export_pdf_fallback(title: str, sections: list[dict[str, Any]]) -> bytes:
    """Fallback PDF generation using basic HTML + weasyprint or simple markup."""
    html = f"<html><head><meta charset='utf-8'><title>{title}</title>"
    html += "<style>body{font-family:sans-serif;padding:20px;}"
    html += "h1{color:#0F5132;}h2{color:#145A32;margin-top:20px;}"
    html += "table{width:100%;border-collapse:collapse;margin:10px 0;}"
    html += "th,td{border:1px solid #ddd;padding:6px;text-align:left;}"
    html += "th{background:#0F5132;color:#fff;}"
    html += "</style></head><body>"
    html += f"<h1>{title}</h1>"

    for section in sections:
        html += f"<h2>{section.get('title', '')}</h2>"
        data = section.get("data", [])
        if data and isinstance(data, list) and data and isinstance(data[0], dict):
            headers = list(data[0].keys())
            html += "<table><thead><tr>" + "".join(f"<th>{h.capitalize()}</th>" for h in headers) + "</tr></thead><tbody>"
            for row in data:
                html += "<tr>" + "".join(f"<td>{row.get(h, '')}</td>" for h in headers) + "</tr>"
            html += "</tbody></table>"
        elif data:
            html += f"<p>{data}</p>"
    html += "</body></html>"

    try:
        import weasyprint
        return weasyprint.HTML(string=html).write_pdf()
    except ImportError:
        pass

    # Absolute last resort: return HTML that can be saved as .html
    return html.encode("utf-8")
